"""
Módulo de generación de reportes Excel con formato
"""
import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.backend.processors.extractor import DataExtractor
from src.backend.processors.transformer import TrayectoriaTransformer
from src.utils.config import config
from src.utils.logger import get_logger


logger = get_logger(__name__, config.paths.logs_dir)


class ExcelGenerator:
    """Generador de reportes Excel con formato"""

    # Colores corporativos
    COLOR_HEADER = "366092"  # Azul oscuro
    COLOR_SUBHEADER = "5B9BD5"  # Azul claro
    COLOR_ALTERNATE = "D9E1F2"  # Gris azulado claro
    COLOR_TOTAL = "FFC000"  # Naranja

    def __init__(self, output_dir: Optional[Path] = None):
        """
        Inicializa el generador

        Args:
            output_dir: Directorio de salida (por defecto config.paths.reports_dir)
        """
        self.output_dir = output_dir or config.paths.reports_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"ExcelGenerator inicializado con output_dir: {self.output_dir}")

    def _apply_header_style(self, ws, row: int, start_col: int, end_col: int):
        """Aplica estilo a la fila de encabezado"""
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color=self.COLOR_HEADER, end_color=self.COLOR_HEADER, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    def _apply_data_style(self, ws, start_row: int, end_row: int, start_col: int, end_col: int):
        """Aplica estilo a las filas de datos"""
        for row in range(start_row, end_row + 1):
            # Alternar colores de fila
            fill_color = self.COLOR_ALTERNATE if row % 2 == 0 else "FFFFFF"

            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.alignment = Alignment(horizontal="left", vertical="center")

    def _auto_adjust_columns(self, ws, min_width: int = 10, max_width: int = 50):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _freeze_panes(self, ws, row: int = 2, col: int = 1):
        """Congela paneles para mantener encabezados visibles"""
        ws.freeze_panes = ws.cell(row=row, column=col)

    def create_worksheet_from_dataframe(
        self,
        wb,
        df: pd.DataFrame,
        sheet_name: str,
        title: Optional[str] = None,
        apply_styles: bool = True
    ):
        """
        Crea una hoja de Excel desde un DataFrame

        Args:
            wb: Workbook de openpyxl
            df: DataFrame con los datos
            sheet_name: Nombre de la hoja
            title: Título opcional sobre la tabla
            apply_styles: Si aplicar estilos automáticos
        """
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)

        current_row = 1

        # Agregar título si se proporciona
        if title:
            ws.cell(row=current_row, column=1, value=title)
            ws.cell(row=current_row, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(df.columns))
            current_row += 2

        # Escribir encabezados
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=current_row, column=col_idx, value=col_name)

        if apply_styles:
            self._apply_header_style(ws, current_row, 1, len(df.columns))

        current_row += 1
        start_data_row = current_row

        # Escribir datos
        for _, row_data in df.iterrows():
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=current_row, column=col_idx, value=value)
            current_row += 1

        if apply_styles and len(df) > 0:
            self._apply_data_style(ws, start_data_row, current_row - 1, 1, len(df.columns))
            self._auto_adjust_columns(ws)
            self._freeze_panes(ws, row=start_data_row, col=1)

        logger.debug(f"Hoja '{sheet_name}' creada con {len(df)} filas")

    def generate_report_for_grado(
        self,
        grado: str,
        year_start: Optional[int] = None,
        year_end: Optional[int] = None,
        filename: Optional[str] = None
    ) -> Path:
        """
        Genera reporte completo para un grado académico

        Args:
            grado: Grado académico (LL, EL, ML)
            year_start: Año inicial (opcional)
            year_end: Año final (opcional)
            filename: Nombre del archivo (opcional, se genera automático)

        Returns:
            Path del archivo generado
        """
        logger.info(f"Generando reporte para {grado}")

        try:
            # Extraer datos
            extractor = DataExtractor(year_start, year_end)
            data = extractor.extract_all_for_grado(grado)

            # Transformar datos
            transformer = TrayectoriaTransformer()

            # Procesar resumen con trayectoria
            df_resumen = transformer.create_trayectoria_table(
                data['nuevo_ingreso'],
                data['reinscritos']
            )

            # Procesar cuadro FIMPES
            df_fimpes = transformer.create_cuadro_fimpes(df_resumen)

            # Nombre del archivo
            if not filename:
                grado_names = {'LL': 'Licenciatura', 'EL': 'Especialidad', 'ML': 'Maestría'}
                timestamp = datetime.now().strftime('%Y%m%d')
                filename = f"Trayectoria_online_{grado_names.get(grado, grado)}_{timestamp}.xlsx"

            output_file = self.output_dir / filename

            # Crear workbook
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Remover hoja por defecto

            # Crear hojas
            self.create_worksheet_from_dataframe(
                wb,
                data['todos'],
                'Hoja1',
                title='Datos Consolidados'
            )

            self.create_worksheet_from_dataframe(
                wb,
                df_resumen,
                'Resumen',
                title=f'Trayectoria {grado}'
            )

            self.create_worksheet_from_dataframe(
                wb,
                data['nuevo_ingreso'],
                'NI',
                title='Nuevo Ingreso'
            )

            self.create_worksheet_from_dataframe(
                wb,
                data['reinscritos'],
                'Reinscritos',
                title='Estudiantes Reinscritos'
            )

            self.create_worksheet_from_dataframe(
                wb,
                df_fimpes,
                'Cuadro FIMPES',
                title='Indicadores FIMPES'
            )

            # Guardar archivo
            wb.save(output_file)
            logger.info(f"Reporte generado exitosamente: {output_file}")

            return output_file

        except Exception as e:
            logger.error(f"Error generando reporte para {grado}: {e}")
            raise

    def generate_all_reports(
        self,
        year_start: Optional[int] = None,
        year_end: Optional[int] = None
    ) -> Dict[str, Path]:
        """
        Genera todos los reportes (LL, EL, ML)

        Args:
            year_start: Año inicial
            year_end: Año final

        Returns:
            Diccionario con los paths de archivos generados
        """
        logger.info("Generando todos los reportes")

        results = {}
        for grado in config.reports.grados:
            try:
                output_file = self.generate_report_for_grado(grado, year_start, year_end)
                results[grado] = output_file
            except Exception as e:
                logger.error(f"Error generando reporte para {grado}: {e}")
                results[grado] = None

        successful = sum(1 for v in results.values() if v is not None)
        logger.info(f"Generación completada: {successful}/{len(results)} reportes exitosos")

        return results


def test_generation():
    """Función de prueba para el generador"""
    logger.info("=== Iniciando test de generación ===")

    try:
        generator = ExcelGenerator()

        # Generar reporte de prueba para LL
        logger.info("\nGenerando reporte de prueba para LL (2024-2025)")
        output_file = generator.generate_report_for_grado('LL', year_start=2024, year_end=2025)

        print(f"\nReporte generado: {output_file}")
        print(f"Tamaño: {output_file.stat().st_size / 1024:.2f} KB")

        logger.info("=== Test completado exitosamente ===")

    except Exception as e:
        logger.error(f"Error en test: {e}")
        raise


if __name__ == "__main__":
    test_generation()
