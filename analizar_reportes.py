"""
Script para analizar los reportes Excel generados
"""
import pandas as pd
import openpyxl
from pathlib import Path
import sys

def analizar_reporte(file_path):
    """Analiza un archivo Excel y muestra su estructura y contenido"""

    print("\n" + "="*80)
    print(f"ANALIZANDO: {file_path.name}")
    print("="*80)

    try:
        # Cargar el workbook con openpyxl para información detallada
        wb = openpyxl.load_workbook(file_path, read_only=True)

        print(f"\n📊 Hojas encontradas: {len(wb.sheetnames)}")
        print(f"   {wb.sheetnames}")

        # Analizar cada hoja
        for sheet_name in wb.sheetnames:
            print(f"\n{'─'*80}")
            print(f"📄 HOJA: {sheet_name}")
            print(f"{'─'*80}")

            # Leer con pandas para análisis detallado
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)

                print(f"\n   📏 Dimensiones: {df.shape[0]} filas x {df.shape[1]} columnas")

                # Mostrar columnas
                print(f"\n   📋 Columnas ({len(df.columns)}):")
                for i, col in enumerate(df.columns, 1):
                    print(f"      {i}. {col}")

                # Mostrar tipos de datos
                print(f"\n   🔤 Tipos de datos:")
                tipos = df.dtypes.value_counts()
                for tipo, count in tipos.items():
                    print(f"      {tipo}: {count} columnas")

                # Verificar valores nulos
                nulos = df.isnull().sum()
                if nulos.sum() > 0:
                    print(f"\n   ⚠️  Valores nulos encontrados:")
                    for col, count in nulos[nulos > 0].items():
                        porcentaje = (count / len(df)) * 100
                        print(f"      {col}: {count} ({porcentaje:.1f}%)")
                else:
                    print(f"\n   ✅ Sin valores nulos")

                # Verificar duplicados (si aplica)
                if 'estudiante_id' in df.columns or 'ID' in df.columns:
                    id_col = 'estudiante_id' if 'estudiante_id' in df.columns else 'ID'
                    periodo_col = None

                    # Buscar columna de periodo
                    for col in df.columns:
                        if 'periodo' in col.lower():
                            periodo_col = col
                            break

                    if periodo_col:
                        duplicados = df.duplicated(subset=[id_col, periodo_col], keep=False)
                        if duplicados.sum() > 0:
                            print(f"\n   ⚠️  Duplicados encontrados: {duplicados.sum()} registros")
                        else:
                            print(f"\n   ✅ Sin duplicados ({id_col} + {periodo_col})")

                # Mostrar primeras filas (solo primeras 3 columnas para legibilidad)
                if len(df) > 0:
                    print(f"\n   👀 Primeras 3 filas (primeras 3 columnas):")
                    print(df.iloc[:3, :min(3, len(df.columns))].to_string(index=False))

                    # Para hoja de resumen, mostrar totales
                    if sheet_name == 'Resumen':
                        print(f"\n   📊 Estadísticas de Resumen:")
                        if 'nuevo_ingreso' in df.columns:
                            print(f"      Total Nuevo Ingreso: {df['nuevo_ingreso'].sum()}")
                        if 'egresados' in df.columns:
                            print(f"      Total Egresados: {df['egresados'].sum()}")
                        if 'titulados' in df.columns:
                            print(f"      Total Titulados: {df['titulados'].sum()}")

            except Exception as e:
                print(f"\n   ❌ Error al leer hoja con pandas: {str(e)}")

        wb.close()
        return True

    except Exception as e:
        print(f"\n❌ ERROR al analizar archivo: {str(e)}")
        return False

def main():
    """Función principal"""

    print("\n" + "="*80)
    print("ANÁLISIS COMPLETO DE REPORTES GENERADOS")
    print("="*80)

    # Directorio de reportes
    reportes_dir = Path("data/reportes_generados")

    # Buscar reportes generados hoy
    reportes = sorted(reportes_dir.glob("*20251216.xlsx"))

    if not reportes:
        print("\n❌ No se encontraron reportes generados hoy (20251216)")
        return

    print(f"\n📁 Reportes encontrados: {len(reportes)}")
    for reporte in reportes:
        print(f"   - {reporte.name}")

    # Analizar cada reporte
    resultados = {}
    for reporte in reportes:
        exito = analizar_reporte(reporte)
        resultados[reporte.name] = exito

    # Resumen final
    print("\n" + "="*80)
    print("RESUMEN DEL ANÁLISIS")
    print("="*80)

    for nombre, exito in resultados.items():
        estado = "✅ Exitoso" if exito else "❌ Fallido"
        print(f"{estado}: {nombre}")

    exitosos = sum(1 for v in resultados.values() if v)
    print(f"\nTotal: {exitosos}/{len(resultados)} reportes analizados exitosamente")

if __name__ == "__main__":
    main()
