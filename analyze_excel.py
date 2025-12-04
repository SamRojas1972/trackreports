#!/usr/bin/env python3
"""
Script para analizar la estructura de los archivos Excel de reportes
"""
import openpyxl
import sys
from pathlib import Path

def analyze_excel(file_path):
    """Analiza un archivo Excel y muestra su estructura"""
    print(f"\n{'='*80}")
    print(f"Analizando: {Path(file_path).name}")
    print(f"{'='*80}\n")

    try:
        # Cargar el workbook
        wb = openpyxl.load_workbook(file_path, data_only=False, keep_vba=True)

        print(f"Número de hojas: {len(wb.sheetnames)}")
        print(f"\nHojas: {', '.join(wb.sheetnames)}")

        # Analizar cada hoja
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n{'-'*80}")
            print(f"Hoja: '{sheet_name}'")
            print(f"{'-'*80}")

            # Dimensiones
            max_row = ws.max_row
            max_col = ws.max_column
            print(f"Dimensiones: {max_row} filas x {max_col} columnas")

            # Rango con datos
            if ws.dimensions:
                print(f"Rango de datos: {ws.dimensions}")

            # Buscar fórmulas en las primeras 100 filas
            formulas = []
            for row in range(1, min(100, max_row + 1)):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row, col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        formulas.append({
                            'cell': cell.coordinate,
                            'formula': cell.value[:100]  # Primeros 100 caracteres
                        })

            if formulas:
                print(f"\nFórmulas encontradas (muestra de primeras {min(len(formulas), 10)}):")
                for f in formulas[:10]:
                    print(f"  {f['cell']}: {f['formula']}")
                if len(formulas) > 10:
                    print(f"  ... y {len(formulas) - 10} más")

            # Mostrar las primeras filas de datos
            print(f"\nPrimeras 5 filas (muestra):")
            for row_idx in range(1, min(6, max_row + 1)):
                row_data = []
                for col_idx in range(1, min(10, max_col + 1)):
                    cell = ws.cell(row_idx, col_idx)
                    value = str(cell.value)[:30] if cell.value else ""
                    row_data.append(value)
                print(f"  Fila {row_idx}: {' | '.join(row_data)}")

            # Buscar formato condicional
            if ws.conditional_formatting:
                print(f"\nFormato condicional: {len(ws.conditional_formatting.cf_rules)} reglas")

            # Buscar tablas dinámicas
            if hasattr(ws, '_pivots') and ws._pivots:
                print(f"\nTablas dinámicas: {len(ws._pivots)}")

    except Exception as e:
        print(f"Error al analizar {file_path}: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    reportes_dir = Path("exampleclaude/reportes")

    excel_files = [
        reportes_dir / "Trayectoria online LL.xlsb",
        reportes_dir / "Trayectoria online EL.xlsb",
        reportes_dir / "Trayectoria online ML.xlsb"
    ]

    for excel_file in excel_files:
        if excel_file.exists():
            analyze_excel(excel_file)
        else:
            print(f"Archivo no encontrado: {excel_file}")
